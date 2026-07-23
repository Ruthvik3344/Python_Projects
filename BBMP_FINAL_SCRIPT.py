import os
import csv
import re
import time
import logging
import requests
import urllib3
from datetime import datetime,timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Border, Side

from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from concurrent.futures import ThreadPoolExecutor, as_completed

START_YEAR = "2016-2017"
END_YEAR = "2026-2027"

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# =====================================================
# CONFIGURATION
# =====================================================

BASE_URL = "https://account.bbmpgov.in/PublicView/?l=1"
API_URL = "https://account.bbmpgov.in/PublicView/vss00CvStatusData.php"

OUTPUT_FOLDER = "BBMP_Development_Data"

MAX_WORKERS = 8

PAGE_SIZE = 100

REQUEST_TIMEOUT = 20
WARD_MAPPING = {
    "oo182 Padmanabha Nagar": "193 Padmanabha Nagar",
}

HEADERS = {
    "User-Agent":
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
    "Accept":
    "application/json, text/javascript, */*; q=0.01",
    "X-Requested-With":
    "XMLHttpRequest"
}

# =====================================================
# CREATE OUTPUT FOLDER
# =====================================================

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# =====================================================
# LOGGING
# =====================================================

logging.basicConfig(
    filename=os.path.join(OUTPUT_FOLDER, "Scraping_Log.txt"),
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s"
)

print("=" * 60)
print("BBMP DEVELOPMENT SCRAPER")
print("=" * 60)

# =====================================================
# REQUEST SESSION
# =====================================================

retry = Retry(
    total=5,
    backoff_factor=1,
    status_forcelist=[429,500,502,503,504],
    allowed_methods=["GET"]
)

adapter = HTTPAdapter(max_retries=retry)

session = requests.Session()

session.headers.update(HEADERS)

session.mount("https://", adapter)
session.mount("http://", adapter)

session.get(BASE_URL, verify=False)

# =====================================================
# HTML CLEANER
# =====================================================

def clean_html(text):

    if text is None:
        return ""

    text = str(text)

    text = text.replace("<br/>"," ")
    text = text.replace("<br>"," ")
    text = text.replace("&nbsp;"," ")

    text = re.sub(r"<[^>]*>","",text)

    return " ".join(text.split())
def format_date(date_text):

    if not date_text:
        return ""

    date_text = str(date_text).strip()

    formats = [
        "%m/%d/%Y",
        "%m/%d/%Y %H:%M:%S",
        "%Y-%m-%d",
        "%d-%m-%Y",
        "%d/%m/%Y",
        "%d-%b-%Y",
        "%d-%b-%Y %H:%M:%S",
        "%d-%b-%y",
        "%d-%b-%y %H:%M:%S"
    ]

    for fmt in formats:
        try:
            dt = datetime.strptime(date_text, fmt)
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            continue

    return date_text

def extract_start_end_dates(jobcode_html):

    start_date = ""
    end_date = ""

    if not jobcode_html:
        return start_date, end_date

    match = re.search(
        r"Start\s*:\s*([0-9]{2}-[A-Za-z]{3}-[0-9]{4}).*?End\s*:\s*([0-9]{2}-[A-Za-z]{3}-[0-9]{4})",
        jobcode_html,
        re.IGNORECASE
    )

    if match:
        start_date = datetime.strptime(
            match.group(1),
            "%d-%b-%Y"
        ).strftime("%d/%m/%Y")

        end_date = datetime.strptime(
            match.group(2),
            "%d-%b-%Y"
        ).strftime("%d/%m/%Y")

    return start_date, end_date


# =====================================================
# WINDOWS SAFE FILENAMES
# =====================================================

def safe_filename(name):

    name = re.sub(r'[\\/:*?"<>|]',"_",name)

    name = name.strip()

    return name

# =====================================================
# GET ALL WARDS
# =====================================================

def get_all_wards():

    print("\nLoading ward list...")

    params = {
        "pAction": "LoadCombo",
        "pTableName": "vssmasters.vss20toward"
    }

    response = session.get(
        API_URL,
        params=params,
        timeout=REQUEST_TIMEOUT,
        verify=False
    )

    wards = response.json()

    ward_list = []

    for ward in wards:

        ward_id = str(ward.get("rid", "")).strip()
        ward_name = clean_html(ward.get("rname", "")).strip()

        # --------------------------
        # Current ward
        # Example:
        # 193 Padmanabha Nagar
        # --------------------------
        m = re.match(r"^(\d{3})\s+(.*)", ward_name)

        if m:
            ward_number = int(m.group(1))

            if ward_number < 1 or ward_number > 255:
                continue
            ward_list.append({
                "id": ward_id,
                "number": int(m.group(1)),
                "generation": "current",
                "name": ward_name
            })
            continue

        # --------------------------
        # Old ward
        # Example:
        # oo182 Padmanabha Nagar
        # --------------------------
        m = re.match(r"^oo(\d{3})\s+(.*)", ward_name, re.IGNORECASE)

        if m:
            ward_number = int(m.group(1))

            if ward_number < 1 or ward_number > 255:
                continue
            ward_list.append({
                "id": ward_id,
                "number": int(m.group(1)),
                "generation": "old",
                "name": ward_name
            })
            continue

        # --------------------------
        # Intermediate ward
        # Example:
        # o182 Shanthi Nagar
        # --------------------------
        m = re.match(r"^o(\d{3})\s+(.*)", ward_name, re.IGNORECASE)

        if m:
            ward_number = int(m.group(1))

            if ward_number < 1 or ward_number > 255:
                continue

            ward_list.append({
                "id": ward_id,
                "number": int(m.group(1)),
                "generation": "intermediate",
                "name": ward_name
            })
            continue

    print(f"Found {len(ward_list)} ward entries.")

    logging.info(f"Loaded {len(ward_list)} ward entries.")

    ward_list.sort(key=lambda x: (x["generation"], x["number"]))

    for w in ward_list:
        print(w)

    return ward_list
# =====================================================
# GET FINANCIAL YEARS
# =====================================================

def get_financial_years():

    print("\nLoading financial years...")

    params = {

        "pAction": "LoadFinancialYear",

        "pTableName": "vss.vss00tvfinancialyear"

    }

    response = session.get(
        API_URL,
        params=params,
        timeout=REQUEST_TIMEOUT,
        verify=False
    )

    years = response.json()

    financial_years = []

    for year in years:

        if year.get("rid") == "-1":
            continue

        start = datetime.strptime(
            year["datefrom"],
            "%Y-%m-%d"
        )

        end = start.replace(year=start.year + 1) - timedelta(days=1)

        financial_years.append({

            "id": year["rid"],

            "name": year["rname"],

            "date_from": start.strftime("%d-%b-%Y"),

            "date_to": end.strftime("%d-%b-%Y")

        })

    print(f"Loaded {len(financial_years)} financial years.")

    return financial_years
# =====================================================
# DOWNLOAD ALL PARENT RECORDS OF A WARD
# =====================================================

def get_parent_records(
    ward_id,
    ward_name,
    financial_years
):

    print(f"\nProcessing Ward : {ward_name}")

    logging.info(f"Processing Ward {ward_name}")

    parent_records = []

    seen_wbids = set()
    for financial_year in financial_years:

        page = 0

        last_page_wbids = None

        while True:

            start_index = page * PAGE_SIZE
            end_index = start_index + PAGE_SIZE

            params = {

                "pAction":"LoadPaymentGridData",

                "pCriteria":"",

                "pDateFrom": financial_year["date_from"],

                "pDateTo": financial_year["date_to"],

                "pDateType": "pDT",

                "pOrderBy": "1",

                "pFinancialYearID":-1,

                "pBudgetHeadID":"-1",

                "pDDOIDs":"",

                "pWardIDs":"",

                "pDateFilterYN":"true",

                "pWardID": ward_id,

                "filterscount":"0",

                "groupscount":"0",

                "pagenum": str(page),

                "pagesize": str(PAGE_SIZE),

                "recordstartindex": str(start_index),

                "recordendindex": str(end_index)

            }
            print(
                f"{financial_year['name']} -> "
                f"{financial_year['date_from']} "
                f"to "
                f"{financial_year['date_to']}"
            )

            try:

                response = session.get(
                    API_URL,
                    params=params,
                    timeout=REQUEST_TIMEOUT,
                    verify=False
                )

                data = response.json()
                print(f"\nWard: {ward_name}")
                print(f"Page: {page}")
                # ------------------------------------------
                # Detect repeated pages returned by BBMP
                # ------------------------------------------

                current_page_wbids = tuple(
                    row.get("wbid", "")
                    for row in data
                )

                if current_page_wbids == last_page_wbids:

                    print("      Repeated page detected. Stopping pagination.")

                    logging.warning(
                        f"{ward_name} ({financial_year['name']}): "
                        f"Repeated page detected."
                    )

                    break

                last_page_wbids = current_page_wbids

            except Exception as e:

                logging.error(f"{ward_name} ({financial_year['name']}) : {e}")

                break

            if not isinstance(data, list):

                break

            if len(data) == 0:

                break

            print(f"      Page {page+1} : {len(data)} rows")

            for row in data:

                wbid = row.get("wbid")

                if not wbid:

                    continue

                unique_key = (
                    financial_year["id"],
                    wbid
                )

                if unique_key in seen_wbids:

                    continue

                seen_wbids.add(unique_key)
                row["Financial_Year"] = financial_year["name"]

                row["Ward_ID"] = ward_id

                parent_records.append(row)

            if len(data) < PAGE_SIZE:

                break

            page += 1

            time.sleep(0.20)

    print(f"   Total Parent Records : {len(parent_records)}")

    logging.info(

        f"{ward_name} -> {len(parent_records)} parent records"

    )

    return parent_records
# =====================================================
# FETCH COMPLETE DETAILS OF ONE WORK BILL
# =====================================================

def fetch_complete_row_data(parent_record, ward_id):
    mobile = ""
    work_bill_id = parent_record.get("wbid")
    start_date, end_date = extract_start_end_dates(
        parent_record.get("jobcode", "")
    )

    if not work_bill_id:
        return []

    # --------------------------
    # LEFT PANEL
    # --------------------------

    meta_params = {

        "pAction":"LoadWorksbillDetails",

        "pWorkBillID":str(work_bill_id)

    }

    budget = ""
    division = ""
    contractor = ""
    bill_type = ""
    gross = ""
    deduction = ""
    net = ""
    approval_level = ""

    try:

        res = session.get(
            API_URL,
            params=meta_params,
            timeout=REQUEST_TIMEOUT,
            verify=False
        )

        meta = res.json()

        if isinstance(meta,dict):

            budget = meta.get("budget","")
            division = meta.get("ddoname","")
            contractor = meta.get("contractorname","")
            bill_type = meta.get("billtype","")
            gross = meta.get("gross","")
            deduction = meta.get("deduction","")
            net = meta.get("nett","")
            approval_level = meta.get("approvallevel","")
            mobile = (
            meta.get("contractormobile1")
            or meta.get("contractormobile")
            or meta.get("mobile")
            or ""
            )

except Exception:

        pass

    # --------------------------
    # RIGHT PANEL
    # --------------------------

    child_params = {

        "pAction":"LoadGridApprovalLevels",

        "pWorkBillID":str(work_bill_id),

        "filterscount":"0",

        "groupscount":"0",

        "pagenum":"0",

        "pagesize":"50",

        "recordstartindex":"0",

        "recordendindex":"50"

    }

    rows = []

    try:

        res = session.get(
            API_URL,
            params=child_params,
            timeout=REQUEST_TIMEOUT,
            verify=False
        )

        approvals = res.json()

    except Exception:

        approvals = []

    if isinstance(approvals,list) and approvals:

        for item in approvals:

            rows.append({

                "Ward_Query_ID/RID":ward_id,
                "Start_Date": format_date(clean_html(meta.get("sbrdate", ""))),
                "End_Date": format_date(clean_html(meta.get("dbrdate", ""))),

                "Mobile": clean_html(mobile),

                "Ward_Generation": parent_record.get("rname", ""),

                "Financial_Year": parent_record.get("Financial_Year", ""),

                "Main_Serial_No":parent_record.get("slno",""),

                "Work_Code_Name":parent_record.get("wcname",""),

                "Work_Bill_ID":work_bill_id,

                "Name_of_Work":clean_html(parent_record.get("nameofwork","")),

                "Total_Amount":parent_record.get("amount",""),

                "Budget_Head":clean_html(budget),

                "Division":clean_html(division),

                "Contractor_Name":clean_html(contractor),

                "Bill_Type":clean_html(bill_type),

                "Gross_Amount":clean_html(gross),

                "Deductions":clean_html(deduction),

                "Net_Amount":clean_html(net),

                "Current_Level":clean_html(approval_level),

                "Detail_Row_ID":item.get("id",""),

                "Approval_Stage_Name":clean_html(item.get("name","")),

                "Milestone_Date": format_date(clean_html(item.get("date", ""))),

                "Approval_Remarks":clean_html(item.get("remarks",""))

            })

        return rows

    return [{

        "Ward_Query_ID/RID":ward_id,
        "Start_Date": format_date(clean_html(meta.get("sbrdate", ""))),
        "End_Date": format_date(clean_html(meta.get("dbrdate", ""))),

        "Mobile": clean_html(mobile),

        "Ward_Generation": parent_record.get("rname", ""),

        "Financial_Year": parent_record.get("Financial_Year", ""),

        "Main_Serial_No":parent_record.get("slno",""),

        "Work_Code_Name":parent_record.get("wcname",""),

        "Work_Bill_ID":work_bill_id,

        "Name_of_Work":clean_html(parent_record.get("nameofwork","")),

        "Total_Amount":parent_record.get("amount",""),

        "Budget_Head":clean_html(budget),

        "Division":clean_html(division),

        "Contractor_Name":clean_html(contractor),

        "Bill_Type":clean_html(bill_type),

        "Gross_Amount":clean_html(gross),

        "Deductions":clean_html(deduction),

        "Net_Amount":clean_html(net),

        "Current_Level":clean_html(approval_level),

        "Detail_Row_ID":"",

        "Approval_Stage_Name":"",

        "Milestone_Date":"",

        "Approval_Remarks":""

    }]
# =====================================================
# PROCESS ENTIRE WARD
# =====================================================

def process_ward(ward,financial_years):
    ward_id = ward["id"]

    ward_name = ward["name"]

    parent_records = get_parent_records(
        ward_id,
        ward_name,
        financial_years
    )

    if len(parent_records)==0:

        return None

    print("   Fetching work bill details...")

    output_rows = []

    with ThreadPoolExecutor(
        max_workers=MAX_WORKERS
    ) as executor:

        futures = [

            executor.submit(
                fetch_complete_row_data,
                record,
                ward_id
            )

            for record in parent_records

        ]

        completed = 0

        for future in as_completed(futures):

            completed += 1

            try:

                rows = future.result()

                output_rows.extend(rows)

            except Exception as ex:

                logging.error(str(ex))

            print(

                f"\r   {completed}/{len(parent_records)}",

                end=""

            )

    print()

    def financial_year_key(row):
        fy = row.get("Financial_Year", "")
        try:
            return int(fy.split("-")[0])
        except Exception:
            return 0

    def work_code_key(row):
        code = row.get("Work_Code_Name", "")
        try:
            return tuple(int(x) for x in code.split("-"))
                             except Exception:
            return (999999,)

    output_rows.sort(
    key=lambda row: (
        -financial_year_key(row),                     # Newest FY first
        int(row["Main_Serial_No"])                    # Then serial number
        if str(row["Main_Serial_No"]).isdigit()
        else 999999,
        work_code_key(row)                         # Finally work code
    )
)

    xlsx_name = safe_filename(
        ward["name"].replace(" ", "_")
    ) + ".xlsx"

    xlsx_path = os.path.join(
        OUTPUT_FOLDER,
        xlsx_name
    )

    fieldnames = [

        "Ward_Query_ID/RID",
        "Financial_Year",
        "Start_Date",
        "End_Date",
        "Main_Serial_No",
        "Work_Code_Name",
        "Work_Bill_ID",
        "Name_of_Work",
        "Total_Amount",
        "Budget_Head",
        "Division",
        "Contractor_Name",
        "Mobile",
        "Bill_Type",
        "Gross_Amount",
        "Deductions",
        "Net_Amount",
        "Current_Level",
        "Detail_Row_ID",
        "Approval_Stage_Name",
        "Milestone_Date",
        "Approval_Remarks"

    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Development Data"
    ws.freeze_panes = "A2"

    # Header row
    ws.append(fieldnames)
    ws.row_dimensions[1].height = 28

    # Data rows
    for row in output_rows:

        ws.append([
            row.get(col, "")
            for col in fieldnames
        ])
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(
        bold=True,
        size=11
    )
    for cell in ws[1]:
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
    for column_cells in ws.columns:

        length = max(
            len(str(cell.value)) if cell.value else 0
            for cell in column_cells
        )

        adjusted_width = min(length + 2, 50)

        ws.column_dimensions[
            get_column_letter(column_cells[0].column)
        ].width = adjusted_width
    # Save workbook
    currency_columns = [
    "Total_Amount",
    "Gross_Amount",
    "Deductions",
    "Net_Amount"
    ]

    header_map = {}

    for cell in ws[1]:
        header_map[cell.value] = cell.column

    for col_name in currency_columns:

        col = header_map.get(col_name)

        if col:

            for row in range(2, ws.max_row + 1):

                cell = ws.cell(row=row, column=col)

                try:
                    cell.value = float(cell.value)
                    cell.number_format = '#,##0.00'
                except:
                    pass
    # ---------------------------------------------------
    # Merge rows having same Main_Serial_No
    # ---------------------------------------------------

    merge_columns = [
        "Start_Date",
        "End_Date",
        "Main_Serial_No",
        "Work_Code_Name",
        "Work_Bill_ID",
        "Name_of_Work",
        "Total_Amount",
        "Budget_Head",
        "Division",
        "Contractor_Name",
        "Mobile",
        "Bill_Type",
        "Gross_Amount",
        "Deductions",
        "Net_Amount",
        "Current_Level"
    ]

    header_map = {}

    for cell in ws[1]:
        header_map[cell.value] = cell.column

    start_row = 2

    while start_row <= ws.max_row:

        serial = ws.cell(
            row=start_row,
            column=header_map["Main_Serial_No"]
        ).value

        end_row = start_row

        current_year = ws.cell(
            row=start_row,
            column=header_map["Financial_Year"]
        ).value

        while (
            end_row + 1 <= ws.max_row
            and ws.cell(
                row=end_row + 1,
                column=header_map["Main_Serial_No"]
            ).value == serial
            and ws.cell(
                row=end_row + 1,
                column=header_map["Financial_Year"]
            ).value == current_year
        ):
            end_row += 1

        if end_row > start_row:

            for col_name in merge_columns:

                col = header_map[col_name]

                ws.merge_cells(
                    start_row=start_row,
                    start_column=col,
                    end_row=end_row,
                    end_column=col
                )

                cell = ws.cell(
                    row=start_row,
                    column=col
                )

                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )

        start_row = end_row + 1
    # Center all cells
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell, MergedCell):
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True
                )
    # ---------------------------------------------------
    # Add borders to all cells
    # ---------------------------------------------------

    thin = Side(
        border_style="thin",
        color="000000"
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    for row in ws.iter_rows():
        for cell in row:
            if cell.coordinate in ws.merged_cells:
                continue
            cell.border = border

    # Also apply border to merged cells
    for merged_range in ws.merged_cells.ranges:

        min_col = merged_range.min_col
        max_col = merged_range.max_col
        min_row = merged_range.min_row
        max_row = merged_range.max_row

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).border = border
    wb.save(xlsx_path)
    print("Merged ranges:")
    print(ws.merged_cells.ranges)

    print(f"   Saved : {xlsx_name}")

    return {

        "Ward": ward_name,
        "Parent": len(parent_records),
        "Rows": len(output_rows),
        "Status": "Success"

    }
# =====================================================
# MAIN
# =====================================================

def main():

    start_time = time.time()

    print("\nLoading all wards...")

    wards = get_all_wards()

    financial_years = get_financial_years()

    summary = []

    total_rows = 0

    total_parent = 0

    print("\n")

    print("="*60)

    print(f"TOTAL WARDS FOUND : {len(wards)}")

    print("="*60)

    for index, ward in enumerate(wards, start=1):
        print()

        print("=" * 70)
        print(f"[{index}/{len(wards)}]")
        print(f"Ward ID   : {ward['id']}")
        print(f"Ward Name : {ward['name']}")
        print("=" * 70)

        try:

            print(f"STARTING -> {ward['name']}")
            result = process_ward(ward,financial_years)
            print(f"FINISHED -> {ward['name']}")
            if result:

                summary.append(result)

                total_rows += result["Rows"]

                total_parent += result["Parent"]

            else:

                summary.append({

                    "Ward":ward["name"],

                    "Parent":0,

                    "Rows":0,

                    "Status":"No Data"

                })

        except Exception as ex:

            logging.exception(ex)
            import traceback
            traceback.print_exc()
            print(f"FAILED WARD : {ward['name']}")
            print(f"FAILED : {ex}")

            summary.append({

                "Ward":ward["name"],

                "Parent":0,

                "Rows":0,

                "Status":"Failed"

            })

            print("FAILED")

        print()

        time.sleep(0.5)

    summary_file = os.path.join(

        OUTPUT_FOLDER,

        "Summary.csv"

    )

    with open(

        summary_file,

        "w",

        newline="",

        encoding="utf-8"

    ) as file:

        writer = csv.DictWriter(

            file,

            fieldnames=[

                "Ward",

                "Parent",

                "Rows",

                "Status"

            ]

        )

        writer.writeheader()

        writer.writerows(summary)

    elapsed = time.time() - start_time

    print()

    print("="*70)

    print("SCRAPING COMPLETED")

    print("="*70)

    print(f"Total Wards      : {len(wards)}")

    print(f"Parent Records   : {total_parent}")

    print(f"Output Rows      : {total_rows}")

    print(f"Time Taken       : {elapsed/60:.2f} minutes")

    print()

    print(f"Output Folder : {OUTPUT_FOLDER}")

    print()

    logging.info("Scraping Finished")



if __name__ == "__main__":

    main()
